#!/usr/bin/python3.5
# the report generated using this script will loop through all tasks within the database and update the matrix

import MySQLdb
import os, sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import urllib.request
import time
from shutil import copyfile

# Initial First File Opening
first = True

def gettask():

    # O pen database connection
    db = MySQLdb.connect("localhost","root","**************","semaphore" )

    # prepare a cursor object using cursor() method
    cursor = db.cursor()
    # Query for completed playbooks
    #Select last playbook ran
    sql1 = "select id,end from task order by id desc"
    try:
        cursor.execute(sql1)
        result = cursor.fetchall()
        for row in result:
            task_id = row[0]
            date = row[1]
            name = getmsg(task_id,date,cursor)
        return name
    except:
        print("failed to query for name")

def getmsg(task_id,date,cursor):

    sql21 = "SELECT output FROM task__output WHERE output like '%Playbook Completed%' "
    sql2 = sql21 + "AND task_id=%s" % task_id
    sql31 = "SELECT output FROM task__output WHERE output like '%Playbook ran%' "
    sql3 = sql31 + "AND task_id=%s" % task_id
    try:
        # Execute the SQL command
        cursor.execute(sql2)
        # Fetch all the rows in a list of lists.
        results = cursor.fetchall()
        try:
            for row in results:
            # Now print fetched result
                msg = row[0]
                result=msg.split("#")
                TTP_ID=result[1]
                status=result[2]
                os=result[3]
                updatereport(TTP_ID,status,date,os)
        except:
            print("Failed to query task")
    except:
        print ("Error: unable to fecth data")
   # cursor2 = db.cursor()
    try:
        cursor.execute(sql3)
        result2 = cursor.fetchone()
        msg = result2[0]
        result3 = msg.split("#")
        name = result3[1]
        return name
    except:
        print("failed to query for name")
    # disconnect from server
    db.close()

def updatereport(TTP_ID,status,date,os):
    # Load in the workbook
    global first
    # Load matrix template
    if (first == True):
        try:
            url= "https://github.com/rallyspeed/ansible-mitre/raw/master/report/Matrix-MITRE.xlsx"
            urllib.request.urlretrieve(url, "Updated-Matrix.xlsx")
            wb = load_workbook('./Updated-Matrix.xlsx')
            first = False
        except:
            print ("Failed to donwlod URL")

    # Load updated matrix
    else:
        wb = load_workbook('/var/www/html/matrix.xlsx')
    if (os == "win"):
        ws = wb['windows']
    elif (os == "linux"):
        ws = wb['linux']
    # ws = wb.active  # Use default/active sheet
    r = int(ws.max_row)
    c = int(ws.max_column)
    try:
        for i in range(1, r+1):
            for j in range(1, c+1):
                cv = ws.cell(row=i, column=j).value
                if TTP_ID == cv:
                # Change color of TTP cell which playbook was run succesfully and add date
                    ws.cell(row=i,column=j).fill = PatternFill(start_color="8470FF", end_color="8470FF",fill_type = "solid")
                    ws.cell(row=i,column=j+1).fill = PatternFill(start_color="8470FF", end_color="8470FF",fill_type = "solid")
                    ws.cell(row=i,column=j+1).value = date
                # Green color 00FF00 if TTP blocked
                    if status == "blocked":
                        ws.cell(row=i,column=j-1).fill = PatternFill(start_color="00FF00", end_color="FF0000",fill_type = "solid")
                # Red color FF0000 if TTP not blocked
                    elif status == "passed":
                        ws.cell(row=i,column=j-1).fill = PatternFill(start_color="FF0000", end_color="00FF00",fill_type = "solid")
                 # Orange color FFA500 if TTP failed
                    elif status == "failed":
                        ws.cell(row=i,column=j-1).fill = PatternFill(start_color="FFA500", end_color="00FF00",fill_type = "solid")
    except:
        print ("failed to loop through xls")
    wb.save("/var/www/html/matrix.xlsx")

def savereport(name):
    t = time.localtime()
    timestamp = time.strftime('%b-%d-%Y_%H%M', t)
    try:
        copyfile("/var/www/html/matrix.xlsx","/var/www/html/matrix-" + name + "-" + timestamp + ".xlsx")
        print("File available: http://*******/matrix-" + name + "-" + timestamp + ".xlsx")
    except:
        print("Failed to copy file")

def main():
    naming=gettask()
    savereport(naming)

if __name__ == '__main__':
    main()
